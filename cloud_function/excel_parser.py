import datetime
from openpyxl import load_workbook

class ExcelParser:
    def __init__(self, file_stream, file_name=None):
        """
        Inicializa o parser com o stream de bytes do arquivo (XLSX, XLS ou CSV).
        """
        self.rows = []
        extension = "xlsx"  # fallback padrão se não puder determinar
        if file_name:
            ext_part = file_name.lower().split('.')[-1]
            if ext_part in ["xlsx", "xls", "csv"]:
                extension = ext_part

        if extension == "xlsx":
            self._load_xlsx(file_stream)
        elif extension == "xls":
            self._load_xls(file_stream)
        elif extension == "csv":
            self._load_csv(file_stream)

        self.transactions = self._parse_sheet()

    def _load_xlsx(self, file_stream):
        wb = load_workbook(file_stream, data_only=True)
        sheet = wb.active
        for row in sheet.iter_rows(values_only=True):
            self.rows.append(list(row))

    def _load_xls(self, file_stream):
        import xlrd
        wb = xlrd.open_workbook(file_contents=file_stream.read())
        sheet = wb.sheet_by_index(0)
        
        for r in range(sheet.nrows):
            row_values = []
            for c in range(sheet.ncols):
                cell = sheet.cell(r, c)
                # Converte datas do Excel (formato numérico) para datetime.date
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        dt_tuple = xlrd.xldate_as_tuple(cell.value, wb.datemode)
                        val = datetime.date(dt_tuple[0], dt_tuple[1], dt_tuple[2])
                    except Exception:
                        val = cell.value
                else:
                    val = cell.value
                row_values.append(val)
            self.rows.append(row_values)

    def _load_csv(self, file_stream):
        import csv
        content = file_stream.read()
        try:
            decoded = content.decode('utf-8')
        except UnicodeDecodeError:
            decoded = content.decode('latin-1')
            
        # Detecta o delimitador (, ou ;) no primeiro bloco de caracteres
        delimiter = ';' if ';' in decoded[:1000] else ','
        
        reader = csv.reader(decoded.splitlines(), delimiter=delimiter)
        for row in reader:
            self.rows.append(row)

    def _parse_sheet(self):
        """
        Varre a lista de linhas unificada, identifica as colunas pelo cabeçalho
        e retorna uma lista de dicionários representando as transações.
        """
        transactions = []
        header_mapping = {}
        
        # Mapeamento esperado das colunas
        expected_columns = {
            "data": ["data"],
            "lancamento": ["lançamento", "lancamento"],
            "dcto": ["dcto.", "dcto", "documento"],
            "credito": ["crédito (r$)", "credito (r$)", "crédito", "credito"],
            "debito": ["débito (r$)", "debito (r$)", "débito", "debito"],
            "saldo": ["saldo (r$)", "saldo", "saldo (r$)"]
        }

        header_row_index = None

        # Procura o cabeçalho nas primeiras 15 linhas (para pular possíveis cabeçalhos de banco)
        for r in range(min(15, len(self.rows))):
            row_values = [str(cell).strip().lower() if cell is not None else "" for cell in self.rows[r]]
            
            # Se acharmos pelo menos 4 colunas esperadas na mesma linha, consideramos como o cabeçalho
            matches = 0
            temp_mapping = {}
            for col_key, aliases in expected_columns.items():
                for idx, val in enumerate(row_values):
                    if val in aliases:
                        temp_mapping[col_key] = idx
                        matches += 1
                        break
            
            if matches >= 4:  # Achou a maioria das colunas principais
                header_mapping = temp_mapping
                header_row_index = r
                break

        if header_row_index is None:
            # Fallback padrão se não encontrar: assume a primeira linha e as colunas padrão
            header_mapping = {
                "data": 0, "lancamento": 1, "dcto": 2, "credito": 3, "debito": 4, "saldo": 5
            }
            header_row_index = 0

        # Iterar pelas linhas de transações abaixo do cabeçalho
        for row in self.rows[header_row_index + 1:]:
            # Ignora linhas vazias ou onde a coluna Data está vazia
            if not row or header_mapping["data"] >= len(row):
                continue
                
            data_val = row[header_mapping["data"]]
            if data_val is None or str(data_val).strip() == "":
                continue

            trans_date = self._parse_date(data_val)
            if not trans_date:
                continue  # Pula se a data for inválida (pode ser uma linha de rodapé)

            trans_time = self._parse_time(data_val)

            # Mapear lançamento
            lanc = str(row[header_mapping["lancamento"]]).strip() if "lancamento" in header_mapping and header_mapping["lancamento"] < len(row) and row[header_mapping["lancamento"]] is not None else ""
            
            # Mapear dcto
            dcto = str(row[header_mapping["dcto"]]).strip() if "dcto" in header_mapping and header_mapping["dcto"] < len(row) and row[header_mapping["dcto"]] is not None else ""

            # Mapear crédito, débito e saldo
            credito = self._parse_float(row[header_mapping["credito"]]) if "credito" in header_mapping and header_mapping["credito"] < len(row) else 0.0
            debito = self._parse_float(row[header_mapping["debito"]]) if "debito" in header_mapping and header_mapping["debito"] < len(row) else 0.0
            saldo = self._parse_float(row[header_mapping["saldo"]]) if "saldo" in header_mapping and header_mapping["saldo"] < len(row) else None

            transactions.append({
                "data": trans_date,
                "hora": trans_time,
                "lancamento": lanc,
                "dcto": dcto,
                "credito": credito,
                "debito": debito,
                "saldo": saldo
            })

        # Ordenar transações por data ascendente para garantir a integridade do saldo acumulado
        transactions.sort(key=lambda x: x["data"])
        return transactions

    def _parse_date(self, val):
        if isinstance(val, datetime.datetime):
            return val.date()
        if isinstance(val, datetime.date):
            return val
        
        # Tenta fazer o parse de string
        val_str = str(val).strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%y", "%d-%m-%Y"):
            try:
                # Divide por espaço caso venha data e hora na string
                date_part = val_str.split(' ')[0].split('T')[0]
                return datetime.datetime.strptime(date_part, fmt).date()
            except ValueError:
                continue
        return None

    def _parse_time(self, val):
        if isinstance(val, datetime.datetime):
            return val.time()
        
        val_str = str(val).strip()
        # Se contiver apenas a data (sem espaço ou T), não tem hora
        if " " not in val_str and "T" not in val_str:
            return None
            
        for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M:%S"):
            try:
                # Tenta extrair a parte de hora ou fazer o parse completo
                return datetime.datetime.strptime(val_str, fmt).time()
            except ValueError:
                continue
        return None

    def _parse_float(self, val):
        if val is None:
            return 0.0
        if isinstance(val, (int, float)):
            return float(val)
        
        val_str = str(val).strip().replace("R$", "").replace(" ", "")
        
        # Verifica se usa vírgula como separador decimal (ex: 1.000,50 ou 1000,50)
        if "," in val_str:
            # Caso tenha ponto de milhar e vírgula decimal (ex: 1.000,50)
            if "." in val_str:
                val_str = val_str.replace(".", "")
            val_str = val_str.replace(",", ".")
            
        try:
            return float(val_str)
        except ValueError:
            return 0.0

    def get_brasilia_today(self):
        """
        Retorna a data de hoje no fuso horário do Brasil (UTC-3).
        """
        utc_now = datetime.datetime.now(datetime.timezone.utc)
        brasilia_offset = datetime.timezone(datetime.timedelta(hours=-3))
        return utc_now.astimezone(brasilia_offset).date()

    def get_today_pix_received(self):
        """
        Retorna a lista de PIX recebidos na data de hoje.
        """
        today = self.get_brasilia_today()
        pix_received = []
        for t in self.transactions:
            if t["data"] == today and t["credito"] > 0:
                desc = t["lancamento"].upper()
                if "PIX" in desc:
                    pix_received.append({
                        "valor": t["credito"],
                        "descricao": t["lancamento"]
                    })
        return pix_received

    def get_balance(self):
        """
        Retorna o saldo mais recente disponível no extrato.
        """
        # Procura a última transação da lista (ordenada por data ascendente) que tenha saldo preenchido
        for t in reversed(self.transactions):
            if t["saldo"] is not None:
                return t["saldo"]
        return 0.0

    def get_yesterday_pix_summary(self):
        """
        Retorna a contagem e valores sumarizados de PIX (recebidos e enviados) ontem.
        """
        today = self.get_brasilia_today()
        yesterday = today - datetime.timedelta(days=1)
        
        count_received = 0
        total_received = 0.0
        count_sent = 0
        total_sent = 0.0

        for t in self.transactions:
            if t["data"] == yesterday:
                desc = t["lancamento"].upper()
                if "PIX" in desc:
                    if t["credito"] > 0:
                        count_received += 1
                        total_received += t["credito"]
                    elif t["debito"] > 0:
                        count_sent += 1
                        total_sent += t["debito"]
                    elif t["debito"] < 0: # Trata débito negativo caso o extrato salve assim
                        count_sent += 1
                        total_sent += abs(t["debito"])

        return {
            "count_received": count_received,
            "total_received": total_received,
            "count_sent": count_sent,
            "total_sent": total_sent,
            "total_count": count_received + count_sent
        }

    def find_pix_sender(self, value):
        """
        Procura por transações recentes de PIX recebido (crédito) com o valor especificado.
        Retorna a descrição/nome do pagador se encontrado.
        """
        target_value = float(value)
        # Varre de trás para frente para pegar o mais recente se houver duplicatas
        for t in reversed(self.transactions):
            if abs(t["credito"] - target_value) < 0.01:
                desc = t["lancamento"].upper()
                if "PIX" in desc:
                    orig_desc = t["lancamento"]
                    cleaned = orig_desc
                    
                    # Remove prefixos comuns
                    prefixes = [
                        "PIX RECEBIDO DE ", "PIX RECEBIDO - ", "PIX RECEBIDO: ", "PIX RECEBIDO ",
                        "PIX TRANSF DE ", "PIX TRANSF - ", "PIX TRANSF ",
                        "PIX CR DE ", "PIX CR ", "PIX CREDITO DE ", "PIX CREDITO ",
                        "PIX RECEB ", "PIX REC "
                    ]
                    
                    for p in prefixes:
                        if cleaned.upper().startswith(p):
                            cleaned = cleaned[len(p):].strip()
                            break
                            
                    return {
                        "valor": t["credito"],
                        "descricao_original": orig_desc,
                        "remetente": cleaned,
                        "data": t["data"].strftime("%d/%m/%Y")
                    }
        return None
